#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File Processor Service
Handles different file types: PDF, Excel, Images
"""

import os
import uuid
import tempfile
from pathlib import Path
from typing import List, Dict, Optional
from PIL import Image

try:
    from pdf2image import convert_from_path
except ImportError:
    convert_from_path = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

import config


class FileProcessor:
    """Process uploaded files and convert to images."""

    def __init__(self):
        self.temp_dir = Path(config.UPLOAD_FOLDER) / "temp"
        self.temp_dir.mkdir(parents=True, exist_ok=True)

    def process_file(self, file_path: str) -> List[str]:
        """
        Process a file and return list of image paths.

        Args:
            file_path: Path to the uploaded file

        Returns:
            List of paths to processed images
        """
        file_type = self.detect_file_type(file_path)

        if file_type == 'pdf':
            return self.pdf_to_images(file_path)
        elif file_type in ['xlsx', 'xls']:
            return self.excel_to_images(file_path)
        elif file_type in ['png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp']:
            return self.process_image(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    def detect_file_type(self, file_path: str) -> str:
        """Detect file type from extension."""
        ext = Path(file_path).suffix.lower().lstrip('.')
        return ext

    def pdf_to_images(self, pdf_path: str) -> List[str]:
        """
        Convert PDF to images.

        Args:
            pdf_path: Path to PDF file

        Returns:
            List of paths to converted images
        """
        if convert_from_path is None:
            raise ImportError("pdf2image not installed. Run: pip install pdf2image")

        # Create unique output directory
        output_dir = self.temp_dir / str(uuid.uuid4())
        output_dir.mkdir(parents=True, exist_ok=True)

        try:
            # Convert PDF pages to images
            # Try different poppler paths for Windows
            poppler_paths = [
                None,  # Default (if poppler in PATH)
                r"C:\Program Files\poppler\Library\bin",
                r"C:\poppler\Library\bin",
                r"C:\Program Files (x86)\poppler\Library\bin",
            ]

            images = None
            for poppler_path in poppler_paths:
                try:
                    if poppler_path and os.path.exists(poppler_path):
                        images = convert_from_path(pdf_path, dpi=200, poppler_path=poppler_path)
                    else:
                        images = convert_from_path(pdf_path, dpi=200)
                    if images:
                        break
                except Exception:
                    continue

            if images is None:
                raise RuntimeError("Could not convert PDF. Please ensure poppler is installed.")

            # Save images
            image_paths = []
            for i, image in enumerate(images):
                image_path = output_dir / f"page_{i+1}.png"
                image.save(str(image_path), 'PNG')
                image_paths.append(str(image_path))

            return image_paths

        except Exception as e:
            raise RuntimeError(f"PDF conversion failed: {str(e)}")

    def excel_to_images(self, excel_path: str) -> List[str]:
        """
        Convert Excel to images by rendering each sheet.

        Args:
            excel_path: Path to Excel file

        Returns:
            List of paths to converted images
        """
        if openpyxl is None:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        # Create unique output directory
        output_dir = self.temp_dir / str(uuid.uuid4())
        output_dir.mkdir(parents=True, exist_ok=True)

        try:
            workbook = openpyxl.load_workbook(excel_path)
            image_paths = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Calculate dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column

                if max_row == 0 or max_col == 0:
                    continue

                # Create image from sheet content
                # Use a simple approach: render as text image
                cell_width = 100
                cell_height = 30
                padding = 20

                img_width = max_col * cell_width + padding * 2
                img_height = max_row * cell_height + padding * 2

                # Create blank image
                img = Image.new('RGB', (img_width, img_height), 'white')

                # We'll use a different approach - save as HTML then convert
                # For now, just create a placeholder with sheet info
                from PIL import ImageDraw, ImageFont

                draw = ImageDraw.Draw(img)

                # Try to use a font
                try:
                    font = ImageFont.truetype("arial.ttf", 14)
                except:
                    font = ImageFont.load_default()

                # Draw header
                draw.text((padding, padding), f"Sheet: {sheet_name}", fill='black', font=font)

                # Draw cells
                y_offset = padding + cell_height
                for row_idx, row in enumerate(sheet.iter_rows(max_row=min(max_row, 50))):
                    x_offset = padding
                    for col_idx, cell in enumerate(row[:min(max_col, 10)]):
                        value = str(cell.value) if cell.value else ''
                        # Truncate long values
                        if len(value) > 15:
                            value = value[:15] + '...'
                        draw.text((x_offset, y_offset), value, fill='black', font=font)
                        x_offset += cell_width
                    y_offset += cell_height

                # Save image
                image_path = output_dir / f"{sheet_name}.png"
                img.save(str(image_path), 'PNG')
                image_paths.append(str(image_path))

            return image_paths

        except Exception as e:
            raise RuntimeError(f"Excel conversion failed: {str(e)}")

    def process_image(self, image_path: str) -> List[str]:
        """
        Process an image file (rotate if needed, optimize).

        Args:
            image_path: Path to image file

        Returns:
            List containing the processed image path
        """
        try:
            img = Image.open(image_path)

            # Auto-rotate based on EXIF data
            if hasattr(img, '_getexif'):
                try:
                    exif = img._getexif()
                    if exif:
                        orientation_key = 274  # EXIF Orientation tag
                        if orientation_key in exif:
                            orientation = exif[orientation_key]
                            if orientation == 2:
                                img = img.transpose(Image.FLIP_LEFT_RIGHT)
                            elif orientation == 3:
                                img = img.rotate(180)
                            elif orientation == 4:
                                img = img.transpose(Image.FLIP_TOP_BOTTOM)
                            elif orientation == 5:
                                img = img.rotate(-90).transpose(Image.FLIP_LEFT_RIGHT)
                            elif orientation == 6:
                                img = img.rotate(-90)
                            elif orientation == 7:
                                img = img.rotate(90).transpose(Image.FLIP_LEFT_RIGHT)
                            elif orientation == 8:
                                img = img.rotate(90)
                except:
                    pass

            # Convert to RGB if necessary
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')

            # Save optimized image
            output_dir = self.temp_dir / str(uuid.uuid4())
            output_dir.mkdir(parents=True, exist_ok=True)

            output_path = output_dir / f"processed_{Path(image_path).name}"
            img.save(str(output_path), 'PNG', optimize=True)

            return [str(output_path)]

        except Exception as e:
            # If processing fails, return original path
            return [image_path]

    def cleanup(self):
        """Clean up temporary files."""
        import shutil
        if self.temp_dir.exists():
            for item in self.temp_dir.iterdir():
                if item.is_dir():
                    shutil.rmtree(item)